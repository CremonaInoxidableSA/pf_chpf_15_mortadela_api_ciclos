import logging
import os
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from services.db import get_connection

logger = logging.getLogger("uvicorn")


def _get_cursor():
    """Obtiene cursor de BD con diccionario como retorno."""
    conn = get_connection()
    return conn, conn.cursor(dictionary=True)


def _cerrar_cursor(conn, cursor):
    """Cierra cursor y conexión."""
    try:
        cursor.close()
        conn.close()
    except Exception:
        pass


def _segundos_a_hhmmss(segundos: int) -> str:
    """Convierte segundos a formato HH:MM:SS."""
    if not segundos:
        return "00:00:00"
    
    horas = segundos // 3600
    minutos = (segundos % 3600) // 60
    segs = segundos % 60
    return f"{horas:02d}:{minutos:02d}:{segs:02d}"


def obtener_recetas_con_ciclos(fecha_inicio: str, fecha_fin: str) -> list:
    """
    Obtiene todas las recetas que tienen ciclos entre las fechas.
    
    Args:
        fecha_inicio: Formato "YYYY-MM-DD"
        fecha_fin: Formato "YYYY-MM-DD"
    
    Returns:
        Lista de id_receta únicos
    """
    conn, cursor = _get_cursor()
    try:
        cursor.execute(
            """
            SELECT DISTINCT c.id_receta
            FROM ciclos c
            WHERE DATE(c.fecha_fin) BETWEEN %s AND %s
            AND c.fecha_fin IS NOT NULL
            ORDER BY c.id_receta
            """,
            (fecha_inicio, fecha_fin)
        )
        result = cursor.fetchall()
        return [row["id_receta"] for row in result]
    finally:
        _cerrar_cursor(conn, cursor)


def obtener_datos_receta(id_receta: int) -> dict:
    """
    Obtiene datos de la receta.
    
    Args:
        id_receta: ID de la receta
    
    Returns:
        Dict con campos de la receta
    """
    conn, cursor = _get_cursor()
    try:
        cursor.execute(
            """
            SELECT id_receta, codigo_producto, peso_producto, 
                   productos_fila, productos_columna
            FROM recetas
            WHERE id_receta = %s
            """,
            (id_receta,)
        )
        result = cursor.fetchone()
        return result if result else {}
    finally:
        _cerrar_cursor(conn, cursor)


def obtener_metricas_receta(id_receta: int, fecha_inicio: str, fecha_fin: str) -> dict:
    """
    Obtiene todas las métricas para una receta en el rango de fechas.
    
    Args:
        id_receta: ID de la receta
        fecha_inicio: Formato "YYYY-MM-DD"
        fecha_fin: Formato "YYYY-MM-DD"
    
    Returns:
        Dict con todas las métricas calculadas
    """
    conn, cursor = _get_cursor()
    try:
        # Obtener ciclos y sus datos
        cursor.execute(
            """
            SELECT c.id_ciclo, c.estado_ciclo, c.tiempo_util
            FROM ciclos c
            WHERE c.id_receta = %s
            AND DATE(c.fecha_fin) BETWEEN %s AND %s
            AND c.fecha_fin IS NOT NULL
            """,
            (id_receta, fecha_inicio, fecha_fin)
        )
        ciclos = cursor.fetchall()
        
        # Obtener todos los niveles de esos ciclos
        if ciclos:
            ciclo_ids = tuple(ciclo["id_ciclo"] for ciclo in ciclos)
            placeholders = ", ".join(["%s"] * len(ciclo_ids))
            
            cursor.execute(
                f"""
                SELECT n.id_ciclo, n.estado_nivel, n.tiempo_nivel, 
                       n.seleccionado, n.finalizado, n.cancelado
                FROM nivelesciclos n
                WHERE n.id_ciclo IN ({placeholders})
                """,
                ciclo_ids
            )
            niveles = cursor.fetchall()
        else:
            niveles = []
        
        # Calcular métricas
        ciclos_totales = len(ciclos)
        ciclos_finalizados = sum(1 for c in ciclos if c["estado_ciclo"] == "FINALIZADO")
        
        niveles_seleccionados = sum(1 for n in niveles if n["seleccionado"])
        niveles_finalizados = sum(1 for n in niveles if n["estado_nivel"] == "FINALIZADO")
        
        # Tiempo útil total (suma de tiempo_nivel en segundos)
        tiempo_util_total = sum(int(n["tiempo_nivel"]) for n in niveles)
        
        # Segundos por nivel: suma tiempo / (finalizados + cancelados)
        niveles_procesados = sum(
            1 for n in niveles 
            if n["estado_nivel"] in ["FINALIZADO", "CANCELADO", "FINALIZADO CON CANCELACIONES"]
        )
        segundos_por_nivel = tiempo_util_total / niveles_procesados if niveles_procesados > 0 else 0
        
        # Porcentaje de falla
        ciclos_cancelados = sum(1 for c in ciclos if c["estado_ciclo"] == "CANCELADO")
        pct_falla = (ciclos_cancelados * 100 / ciclos_totales) if ciclos_totales > 0 else 0
        
        # Eficiencia bruta por nivel
        # Numerador: niveles FINALIZADO, CANCELADO o FINALIZADO CON CANCELACIONES (menos cancelados)
        niveles_exitosos = sum(
            1 for n in niveles 
            if n["estado_nivel"] == "FINALIZADO"
        )
        
        niveles_con_falla = sum(
            1 for n in niveles 
            if n["estado_nivel"] in ["CANCELADO", "FINALIZADO CON CANCELACIONES"]
        )
        
        eficiencia_bruta = (
            (niveles_exitosos * 100 / (niveles_exitosos + niveles_con_falla)) 
            if (niveles_exitosos + niveles_con_falla) > 0 else 0
        )
        
        return {
            "id_receta": id_receta,
            "ciclos_totales": ciclos_totales,
            "ciclos_finalizados": ciclos_finalizados,
            "ciclos_cancelados": ciclos_cancelados,
            "niveles_seleccionados": niveles_seleccionados,
            "niveles_finalizados": niveles_finalizados,
            "niveles_exitosos": niveles_exitosos,
            "niveles_con_falla": niveles_con_falla,
            "tiempo_util_segundos": tiempo_util_total,
            "segundos_por_nivel": segundos_por_nivel,
            "pct_falla": pct_falla,
            "eficiencia_bruta": eficiencia_bruta,
            "niveles_procesados": niveles_procesados,
        }
    finally:
        _cerrar_cursor(conn, cursor)


def calcular_kilos_por_hora(id_receta: int, fecha_inicio: str, fecha_fin: str, 
                            peso_producto: float, productos_fila: int, productos_columna: int,
                            tiempo_util_segundos: int) -> float:
    """
    Calcula kilos por hora procesados.
    
    Fórmula: (Suma de niveles con finalizado=true * peso_producto * productos_fila * productos_columna) / Tiempo util en horas
    
    Args:
        id_receta: ID de la receta
        fecha_inicio: Formato "YYYY-MM-DD"
        fecha_fin: Formato "YYYY-MM-DD"
        peso_producto: Peso unitario del producto
        productos_fila: Cantidad de productos por fila
        productos_columna: Cantidad de productos por columna
        tiempo_util_segundos: Tiempo útil en segundos
    
    Returns:
        Kilos por hora
    """
    if tiempo_util_segundos <= 0:
        return 0.0
    
    conn, cursor = _get_cursor()
    try:
        # Obtener ciclos
        cursor.execute(
            """
            SELECT c.id_ciclo
            FROM ciclos c
            WHERE c.id_receta = %s
            AND DATE(c.fecha_fin) BETWEEN %s AND %s
            AND c.fecha_fin IS NOT NULL
            """,
            (id_receta, fecha_inicio, fecha_fin)
        )
        ciclos = cursor.fetchall()
        
        if not ciclos:
            return 0.0
        
        ciclo_ids = tuple(ciclo["id_ciclo"] for ciclo in ciclos)
        placeholders = ", ".join(["%s"] * len(ciclo_ids))
        
        # Obtener niveles finalizados
        cursor.execute(
            f"""
            SELECT COUNT(*) as cantidad
            FROM nivelesciclos n
            WHERE n.id_ciclo IN ({placeholders})
            AND n.estado_nivel = 'FINALIZADO'
            """,
            ciclo_ids
        )
        result = cursor.fetchone()
        niveles_finalizados = result["cantidad"] if result else 0
        
        # Calcular kilos
        kilos_totales = niveles_finalizados * peso_producto * productos_fila * productos_columna
        
        # Convertir tiempo de segundos a horas
        horas = tiempo_util_segundos / 3600
        
        kilos_por_hora = kilos_totales / horas if horas > 0 else 0
        
        return kilos_por_hora
    finally:
        _cerrar_cursor(conn, cursor)


def obtener_nombre_rack(id_rack: int) -> str:
    """Obtiene el nombre_rack del rack por su ID."""
    conn, cursor = _get_cursor()
    try:
        cursor.execute(
            "SELECT nombre_rack FROM racks WHERE id_rack = %s",
            (id_rack,)
        )
        result = cursor.fetchone()
        return result["nombre_rack"] if result else f"Rack {id_rack}"
    finally:
        _cerrar_cursor(conn, cursor)


def obtener_nombre_receta(id_receta: int) -> str:
    """Obtiene el nombre/código del producto de la receta."""
    conn, cursor = _get_cursor()
    try:
        cursor.execute(
            "SELECT codigo_producto FROM recetas WHERE id_receta = %s",
            (id_receta,)
        )
        result = cursor.fetchone()
        return result["codigo_producto"] if result else f"Receta {id_receta}"
    finally:
        _cerrar_cursor(conn, cursor)


def obtener_ciclos_con_detalles(fecha_inicio: str, fecha_fin: str) -> list:
    """
    Obtiene todos los ciclos con detalles de niveles.
    
    Args:
        fecha_inicio: Formato "YYYY-MM-DD"
        fecha_fin: Formato "YYYY-MM-DD"
    
    Returns:
        Lista de dicts con info de ciclos y sus niveles
    """
    conn, cursor = _get_cursor()
    try:
        cursor.execute(
            """
            SELECT c.id_ciclo, c.id_rack, c.estado_ciclo, c.id_receta,
                   c.tiempo_total, c.tiempo_util, c.fecha_inicio, c.fecha_fin
            FROM ciclos c
            WHERE DATE(c.fecha_fin) BETWEEN %s AND %s
            AND c.fecha_fin IS NOT NULL
            ORDER BY c.fecha_inicio
            """,
            (fecha_inicio, fecha_fin)
        )
        return cursor.fetchall()
    finally:
        _cerrar_cursor(conn, cursor)


def obtener_niveles_ciclo(id_ciclo: int) -> dict:
    """
    Obtiene estadísticas de niveles para un ciclo.
    
    Args:
        id_ciclo: ID del ciclo
    
    Returns:
        Dict con: niveles_seleccionados, niveles_finalizados, tiempo_total_nivel
    """
    conn, cursor = _get_cursor()
    try:
        cursor.execute(
            """
            SELECT 
                SUM(CASE WHEN seleccionado = true THEN 1 ELSE 0 END) as niveles_seleccionados,
                SUM(CASE WHEN finalizado = true THEN 1 ELSE 0 END) as niveles_finalizados,
                SUM(tiempo_nivel) as tiempo_total_nivel,
                COUNT(id_nivel) as niveles_procesados
            FROM nivelesciclos
            WHERE id_ciclo = %s
            """,
            (id_ciclo,)
        )
        result = cursor.fetchone()
        return {
            "niveles_seleccionados": result["niveles_seleccionados"] or 0,
            "niveles_finalizados": result["niveles_finalizados"] or 0,
            "tiempo_total_nivel": result["tiempo_total_nivel"] or 0,
            "niveles_procesados": result["niveles_procesados"] or 0
        }
    finally:
        _cerrar_cursor(conn, cursor)


def obtener_niveles_con_ciclos(fecha_inicio: str, fecha_fin: str) -> list:
    """
    Obtiene todos los niveles de ciclos en el rango de fechas.
    
    Args:
        fecha_inicio: Formato "YYYY-MM-DD"
        fecha_fin: Formato "YYYY-MM-DD"
    
    Returns:
        Lista de dicts con info de niveles y sus ciclos asociados
    """
    conn, cursor = _get_cursor()
    try:
        cursor.execute(
            """
            SELECT nc.id_nivel, nc.id_ciclo, nc.nivel, nc.estado_nivel, nc.tiempo_nivel
            FROM nivelesciclos nc
            JOIN ciclos c ON nc.id_ciclo = c.id_ciclo
            WHERE DATE(c.fecha_fin) BETWEEN %s AND %s
            AND c.fecha_fin IS NOT NULL
            ORDER BY c.fecha_inicio, nc.id_ciclo, nc.nivel
            """,
            (fecha_inicio, fecha_fin)
        )
        return cursor.fetchall()
    finally:
        _cerrar_cursor(conn, cursor)


def obtener_id_rack_por_ciclo(id_ciclo: int) -> int:
    """Obtiene el id_rack asociado a un ciclo."""
    conn, cursor = _get_cursor()
    try:
        cursor.execute(
            "SELECT id_rack FROM ciclos WHERE id_ciclo = %s",
            (id_ciclo,)
        )
        result = cursor.fetchone()
        return result["id_rack"] if result else None
    finally:
        _cerrar_cursor(conn, cursor)


def generar_reporte_excel(fecha_inicio: str, fecha_fin: str) -> BytesIO:
    """
    Genera el archivo Excel con el reporte de ciclos.
    
    Args:
        fecha_inicio: Formato "YYYY-MM-DD"
        fecha_fin: Formato "YYYY-MM-DD"
    
    Returns:
        BytesIO con el archivo Excel
    """
    logger.info(f"Generando reporte de ciclos: {fecha_inicio} a {fecha_fin}")
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "RESUMEN"
        
        # Encabezado principal
        ws.merge_cells("A1:H1")
        ws["A1"] = "RESUMEN DE PRODUCTIVIDAD GENERAL | CREMINOX (ARMADOR/DESARMADOR)"
        ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A1"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws.row_dimensions[1].height = 30
        
        # Fecha de filtrado
        ws["A3"] = "Fecha inicial de filtrado:"
        ws["A3"].font = Font(size=12, bold=True)
        ws["A4"] = "Fecha final de filtrado:"
        ws["A4"].font = Font(size=12, bold=True)
        ws["B3"] = fecha_inicio
        ws["B4"] = fecha_fin
        ws["B3"].font = Font(size=11)
        ws["B4"].font = Font(size=11)
        
        # Insertar logo a la derecha (G3)
        logo_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "static", "cremonarecort.png")
        if os.path.exists(logo_path):
            try:
                img = XLImage(logo_path)
                img.width = 126
                img.height = 31.5
                ws.add_image(img, "H3")
            except Exception as e:
                logger.warning(f"No se pudo insertar logo: {e}")
        
        # Encabezados de tabla
        headers = [
            "Tipo de corte",
            "Cantidad de racks\n(Exitosos/Total)",
            "Cantidad de niveles\n(Selec./Exitosos)",
            "Tiempo útil de proceso\n[HH:MM:SS]",
            "Segundos/Nivel\n[seg]",
            "Kilos por hora\n[kg/h]",
            "% Falla",
            "Eficiencia bruta\npor nivel [%]"
        ]
        
        # Fila 6 para encabezados
        ws.append([])  # Fila vacía
        ws.append(headers)
        
        # Formatear encabezados
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        header_row = ws.max_row
        for cell in ws[header_row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        ws.row_dimensions[header_row].height = 50
        
        # Obtener recetas con ciclos
        recetas_ids = obtener_recetas_con_ciclos(fecha_inicio, fecha_fin)
        
        totales = {
            "ciclos_totales": 0,
            "ciclos_finalizados": 0,
            "ciclos_cancelados": 0,
            "niveles_seleccionados": 0,
            "niveles_finalizados": 0,
            "niveles_exitosos": 0,
            "niveles_con_falla": 0,
            "tiempo_util_segundos": 0,
            "kilos_totales": 0,
            "horas_totales": 0,
            "niveles_procesados_total": 0,
        }
        
        # Definir bordes
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Procesar cada receta
        for id_receta in recetas_ids:
            datos_receta = obtener_datos_receta(id_receta)
            metricas = obtener_metricas_receta(id_receta, fecha_inicio, fecha_fin)
            
            if not datos_receta:
                continue
            
            # Calcular kilos por hora
            kilos_por_hora = calcular_kilos_por_hora(
                id_receta, fecha_inicio, fecha_fin,
                float(datos_receta.get("peso_producto", 0)),
                int(datos_receta.get("productos_fila", 1)),
                int(datos_receta.get("productos_columna", 1)),
                metricas["tiempo_util_segundos"]
            )
            
            # Armar fila
            tipo_corte = datos_receta.get("codigo_producto", "")
            cantidad_racks = f"{metricas['ciclos_finalizados']}/{metricas['ciclos_totales']}"
            cantidad_niveles = f"{metricas['niveles_seleccionados']}/{metricas['niveles_finalizados']}"
            tiempo_util_formato = _segundos_a_hhmmss(metricas["tiempo_util_segundos"])
            segundos_por_nivel_formato = f"{metricas['segundos_por_nivel']:.1f}"
            kilos_por_hora_formato = f"{kilos_por_hora:.2f}"
            pct_falla_formato = f"{metricas['pct_falla']:.1f}"
            eficiencia_bruta_formato = f"{metricas['eficiencia_bruta']:.1f}"
            
            ws.append([
                tipo_corte,
                cantidad_racks,
                cantidad_niveles,
                tiempo_util_formato,
                segundos_por_nivel_formato,
                kilos_por_hora_formato,
                pct_falla_formato,
                eficiencia_bruta_formato
            ])
            
            # Formatear fila de datos
            data_row = ws.max_row
            for col in range(1, 9):
                cell = ws.cell(row=data_row, column=col)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            
            ws.row_dimensions[data_row].height = 25
            
            # Acumular totales
            totales["ciclos_totales"] += metricas["ciclos_totales"]
            totales["ciclos_finalizados"] += metricas["ciclos_finalizados"]
            totales["ciclos_cancelados"] += metricas["ciclos_cancelados"]
            totales["niveles_seleccionados"] += metricas["niveles_seleccionados"]
            totales["niveles_finalizados"] += metricas["niveles_finalizados"]
            totales["niveles_exitosos"] += metricas["niveles_exitosos"]
            totales["niveles_con_falla"] += metricas["niveles_con_falla"]
            totales["tiempo_util_segundos"] += metricas["tiempo_util_segundos"]
            totales["niveles_procesados_total"] += metricas["niveles_procesados"]
            totales["kilos_totales"] += (
                metricas["niveles_finalizados"] *
                float(datos_receta.get("peso_producto", 0)) *
                int(datos_receta.get("productos_fila", 1)) *
                int(datos_receta.get("productos_columna", 1))
            )
        
        # Calcular horas para kilos por hora total
        if totales["tiempo_util_segundos"] > 0:
            totales["horas_totales"] = totales["tiempo_util_segundos"] / 3600
        
        # Fila de totales
        kilos_por_hora_total = totales["kilos_totales"] / totales["horas_totales"] if totales["horas_totales"] > 0 else 0
        
        # Porcentaje de falla total
        pct_falla_total = (
            (totales["ciclos_cancelados"] * 100 / totales["ciclos_totales"])
        ) if totales["ciclos_totales"] > 0 else 0
        
        # Segundos por nivel total
        segundos_por_nivel_total = (
            totales["tiempo_util_segundos"] / totales["niveles_procesados_total"]
        ) if totales["niveles_procesados_total"] > 0 else 0
        
        # Eficiencia bruta total
        eficiencia_bruta_total = (
            (totales["niveles_exitosos"] * 100 / (totales["niveles_exitosos"] + totales["niveles_con_falla"]))
        ) if (totales["niveles_exitosos"] + totales["niveles_con_falla"]) > 0 else 0
        
        tiempo_util_total_formato = _segundos_a_hhmmss(totales["tiempo_util_segundos"])
        
        ws.append([
            "TOTALES",
            f"{totales['ciclos_finalizados']}/{totales['ciclos_totales']}",
            f"{totales['niveles_seleccionados']}/{totales['niveles_finalizados']}",
            tiempo_util_total_formato,
            f"{segundos_por_nivel_total:.1f}",
            f"{kilos_por_hora_total:.2f}",
            f"{pct_falla_total:.1f}",
            f"{eficiencia_bruta_total:.1f}"
        ])
        
        # Formatear fila de totales
        fila_totales = ws.max_row
        totales_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        totales_font = Font(bold=True, size=12)
        
        for col in range(1, 9):
            cell = ws.cell(row=fila_totales, column=col)
            cell.fill = totales_fill
            cell.font = totales_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        
        ws.row_dimensions[fila_totales].height = 30
        
        # Ajustar ancho de columnas
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 25
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 18
        ws.column_dimensions["G"].width = 12
        ws.column_dimensions["H"].width = 20
        
        # ==================== HOJA RACKS ====================
        ws_racks = wb.create_sheet("RACKS")
        
        # Encabezado principal
        ws_racks.merge_cells("A1:J1")
        ws_racks["A1"] = "RESUMEN DE PRODUCTIVIDAD POR RACK | CREMINOX (ARMADOR/DESARMADOR)"
        ws_racks["A1"].font = Font(size=16, bold=True, color="FFFFFF")
        ws_racks["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws_racks["A1"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws_racks.row_dimensions[1].height = 30
        
        # Fecha de filtrado
        ws_racks["A3"] = "Fecha inicial de filtrado:"
        ws_racks["A3"].font = Font(size=12, bold=True)
        ws_racks["A4"] = "Fecha final de filtrado:"
        ws_racks["A4"].font = Font(size=12, bold=True)
        ws_racks["B3"] = fecha_inicio
        ws_racks["B4"] = fecha_fin
        ws_racks["B3"].font = Font(size=11)
        ws_racks["B4"].font = Font(size=11)
        
        # Insertar logo a la derecha (J3)
        if os.path.exists(logo_path):
            try:
                img = XLImage(logo_path)
                img.width = 126
                img.height = 31.5
                ws_racks.add_image(img, "J3")
            except Exception as e:
                logger.warning(f"No se pudo insertar logo en hoja RACKS: {e}")
        
        # Encabezados de columnas
        headers_racks = [
            "ID Ciclo",
            "Rack procesado",
            "Resultado",
            "Tipo de corte",
            "Cantidad de niveles\n(Selec./Exitosos)",
            "Segundos/Nivel\n[seg]",
            "Tiempo total ciclo\n[HH:MM:SS]",
            "Tiempo util ciclo\n[HH:MM:SS]",
            "Inicio\n[AAAA-MM-DD HH:MM:SS]",
            "Fin\n[AAAA-MM-DD HH:MM:SS]"
        ]
        
        ws_racks.append([])  # Fila vacía
        ws_racks.append(headers_racks)
        
        header_row_racks = ws_racks.max_row
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for cell in ws_racks[header_row_racks]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        ws_racks.row_dimensions[header_row_racks].height = 50
        
        # Obtener ciclos
        ciclos = obtener_ciclos_con_detalles(fecha_inicio, fecha_fin)
        
        # Variables para totales
        totales_racks = {
            "niveles_seleccionados_total": 0,
            "niveles_finalizados_total": 0,
            "tiempo_total_nivel_total": 0,
            "tiempo_total_ciclo_total": 0,
            "tiempo_util_ciclo_total": 0,
            "niveles_procesados_total": 0
        }
        
        # Procesar cada ciclo
        for ciclo in ciclos:
            id_ciclo = ciclo["id_ciclo"]
            id_rack = ciclo["id_rack"]
            estado_ciclo = ciclo["estado_ciclo"]
            id_receta = ciclo["id_receta"]
            tiempo_total = ciclo["tiempo_total"] or 0
            tiempo_util = ciclo["tiempo_util"] or 0
            fecha_inicio_ciclo = ciclo["fecha_inicio"]
            fecha_fin_ciclo = ciclo["fecha_fin"]
            
            # Obtener detalles de niveles
            niveles_info = obtener_niveles_ciclo(id_ciclo)
            niveles_seleccionados = niveles_info["niveles_seleccionados"]
            niveles_finalizados = niveles_info["niveles_finalizados"]
            tiempo_total_nivel = niveles_info["tiempo_total_nivel"]
            niveles_procesados = niveles_info["niveles_procesados"]
            
            # Calcular segundos/nivel
            segundos_por_nivel = 0
            if niveles_procesados > 0 and tiempo_total_nivel > 0:
                segundos_por_nivel = tiempo_total_nivel / niveles_procesados
            
            # Obtener nombres
            nombre_rack = obtener_nombre_rack(id_rack)
            nombre_receta = obtener_nombre_receta(id_receta)
            
            # Formatear fechas
            if fecha_inicio_ciclo:
                fecha_inicio_format = fecha_inicio_ciclo.strftime("%Y-%m-%d %H:%M:%S")
            else:
                fecha_inicio_format = ""
            
            if fecha_fin_ciclo:
                fecha_fin_format = fecha_fin_ciclo.strftime("%Y-%m-%d %H:%M:%S")
            else:
                fecha_fin_format = ""
            
            # Formatear tiempos
            tiempo_total_format = _segundos_a_hhmmss(tiempo_total)
            tiempo_util_format = _segundos_a_hhmmss(tiempo_util)
            
            # Agregar fila
            ws_racks.append([
                id_ciclo,
                nombre_rack,
                estado_ciclo,
                nombre_receta,
                f"{niveles_seleccionados}/{niveles_finalizados}",
                f"{segundos_por_nivel:.1f}",
                tiempo_total_format,
                tiempo_util_format,
                fecha_inicio_format,
                fecha_fin_format
            ])
            
            # Acumular totales
            totales_racks["niveles_seleccionados_total"] += niveles_seleccionados
            totales_racks["niveles_finalizados_total"] += niveles_finalizados
            totales_racks["tiempo_total_nivel_total"] += tiempo_total_nivel
            totales_racks["tiempo_total_ciclo_total"] += tiempo_total
            totales_racks["tiempo_util_ciclo_total"] += tiempo_util
            totales_racks["niveles_procesados_total"] += niveles_procesados
        
        # Calcular promedio de segundos/nivel para totales
        segundos_por_nivel_promedio = 0
        if totales_racks["niveles_procesados_total"] > 0 and totales_racks["tiempo_total_nivel_total"] > 0:
            segundos_por_nivel_promedio = totales_racks["tiempo_total_nivel_total"] / totales_racks["niveles_procesados_total"]
        
        # Fila de totales
        tiempo_total_ciclo_total_format = _segundos_a_hhmmss(totales_racks["tiempo_total_ciclo_total"])
        tiempo_util_ciclo_total_format = _segundos_a_hhmmss(totales_racks["tiempo_util_ciclo_total"])
        
        ws_racks.append([
            "",
            "",
            "",
            "",
            f"{totales_racks['niveles_seleccionados_total']}/{totales_racks['niveles_finalizados_total']}",
            f"{segundos_por_nivel_promedio:.1f}",
            tiempo_total_ciclo_total_format,
            tiempo_util_ciclo_total_format,
            "",
            ""
        ])
        
        # Formatear fila de totales
        fila_totales_racks = ws_racks.max_row
        totales_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        totales_font = Font(bold=True, size=12)
        
        for col in range(1, 11):
            cell = ws_racks.cell(row=fila_totales_racks, column=col)
            cell.fill = totales_fill
            cell.font = totales_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        
        ws_racks.row_dimensions[fila_totales_racks].height = 30
        
        # Aplicar formato a datos
        for row_num in range(header_row_racks + 1, fila_totales_racks):
            for col_num in range(1, 11):
                cell = ws_racks.cell(row=row_num, column=col_num)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
        
        # Ajustar ancho de columnas
        ws_racks.column_dimensions["A"].width = 12
        ws_racks.column_dimensions["B"].width = 20
        ws_racks.column_dimensions["C"].width = 20
        ws_racks.column_dimensions["D"].width = 18
        ws_racks.column_dimensions["E"].width = 20
        ws_racks.column_dimensions["F"].width = 15
        ws_racks.column_dimensions["G"].width = 22
        ws_racks.column_dimensions["H"].width = 22
        ws_racks.column_dimensions["I"].width = 28
        ws_racks.column_dimensions["J"].width = 28
        
        # ==================== HOJA NIVELES ====================
        ws_niveles = wb.create_sheet("NIVELES")
        
        # Encabezado principal
        ws_niveles.merge_cells("A1:E1")
        ws_niveles["A1"] = "PRODUCTIVIDAD POR NIVEL | CREMINOX (ARMADOR/DESARMADOR)"
        ws_niveles["A1"].font = Font(size=16, bold=True, color="FFFFFF")
        ws_niveles["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws_niveles["A1"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws_niveles.row_dimensions[1].height = 30
        
        # Fecha de filtrado
        ws_niveles["A3"] = "Fecha inicial de filtrado:"
        ws_niveles["A3"].font = Font(size=12, bold=True)
        ws_niveles["A4"] = "Fecha final de filtrado:"
        ws_niveles["A4"].font = Font(size=12, bold=True)
        ws_niveles["B3"] = fecha_inicio
        ws_niveles["B4"] = fecha_fin
        ws_niveles["B3"].font = Font(size=11)
        ws_niveles["B4"].font = Font(size=11)
        
        # Insertar logo a la derecha (E3)
        if os.path.exists(logo_path):
            try:
                img = XLImage(logo_path)
                img.width = 126
                img.height = 31.5
                ws_niveles.add_image(img, "E3")
            except Exception as e:
                logger.warning(f"No se pudo insertar logo en hoja NIVELES: {e}")
        
        # Encabezados de columnas
        headers_niveles = [
            "ID Ciclo",
            "Rack",
            "Nivel",
            "Resultado",
            "Tiempo útil nivel\n[MM:SS]"
        ]
        
        ws_niveles.append([])  # Fila vacía
        ws_niveles.append(headers_niveles)
        
        header_row_niveles = ws_niveles.max_row
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for cell in ws_niveles[header_row_niveles]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        ws_niveles.row_dimensions[header_row_niveles].height = 50
        
        # Obtener niveles
        niveles = obtener_niveles_con_ciclos(fecha_inicio, fecha_fin)
        
        # Procesar cada nivel
        for nivel in niveles:
            id_ciclo = nivel["id_ciclo"]
            nivel_num = nivel["nivel"]
            estado_nivel = nivel["estado_nivel"]
            tiempo_nivel = nivel["tiempo_nivel"] or 0
            
            # Obtener id_rack del ciclo y luego el nombre del rack
            id_rack = obtener_id_rack_por_ciclo(id_ciclo)
            nombre_rack = obtener_nombre_rack(id_rack) if id_rack else "N/A"
            
            # Formatear tiempo en MM:SS
            minutos = tiempo_nivel // 60
            segundos = tiempo_nivel % 60
            tiempo_nivel_format = f"{minutos:02d}:{segundos:02d}"
            
            # Agregar fila
            ws_niveles.append([
                id_ciclo,
                nombre_rack,
                nivel_num,
                estado_nivel,
                tiempo_nivel_format
            ])
        
        # Aplicar formato a datos
        for row_num in range(header_row_niveles + 1, ws_niveles.max_row + 1):
            for col_num in range(1, 6):
                cell = ws_niveles.cell(row=row_num, column=col_num)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
        
        # Ajustar ancho de columnas
        ws_niveles.column_dimensions["A"].width = 12
        ws_niveles.column_dimensions["B"].width = 20
        ws_niveles.column_dimensions["C"].width = 12
        ws_niveles.column_dimensions["D"].width = 25
        ws_niveles.column_dimensions["E"].width = 20
        
        # Guardar a BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info(f"Reporte generado exitosamente")
        return output
    
    except Exception as e:
        logger.error(f"Error generando reporte: {e}", exc_info=True)
        raise
